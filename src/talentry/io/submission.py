"""Submission CSV writer - strictly conforming to `validate_submission.py`.

The official validator (`validate_submission.py`) enforces:

1. Exactly the header ``candidate_id,rank,score,reasoning``.
2. Exactly 100 data rows.
3. Each ``rank`` 1..100 used once.
4. ``candidate_id`` matches ``^CAND_[0-9]{7}$``.
5. ``score`` non-increasing as rank increases.
6. Equal-score tie-break: ``candidate_id`` ascending.

We re-implement those invariants here so the writer **cannot** produce a CSV
the validator would reject - failing loudly at write time is much friendlier
than discovering the problem at upload time.
"""

from __future__ import annotations

import csv
import re
from collections.abc import Iterable
from pathlib import Path

from talentry.core.models import RankedCandidate

HEADER: tuple[str, ...] = ("candidate_id", "rank", "score", "reasoning")
_CAND_RE = re.compile(r"^CAND_[0-9]{7}$")


class SubmissionError(ValueError):
    """Raised when the writer is asked to emit an invalid submission."""


def _enforce_invariants(rows: list[RankedCandidate]) -> list[RankedCandidate]:
    if len(rows) != 100:
        raise SubmissionError(f"expected exactly 100 rows, got {len(rows)}")

    seen_ids: set[str] = set()
    for r in rows:
        if not _CAND_RE.match(r.candidate_id):
            raise SubmissionError(f"candidate_id '{r.candidate_id}' violates CAND_XXXXXXX format")
        if r.candidate_id in seen_ids:
            raise SubmissionError(f"duplicate candidate_id: {r.candidate_id}")
        seen_ids.add(r.candidate_id)

    # Scores in the CSV are quantised to 4 decimals; what looked like a strict
    # ordering at full precision can collapse into ties after rounding. So we
    # canonically re-sort by (quantised_score DESC, candidate_id ASC) and
    # re-issue ranks 1..100 to guarantee both invariants the official
    # validator checks.
    rows = sorted(rows, key=lambda r: (-round(r.score, 4), r.candidate_id))
    canon: list[RankedCandidate] = []
    for new_rank, r in enumerate(rows, start=1):
        # RankedCandidate is a dataclass - mutate `rank` in place.
        try:
            r.rank = new_rank  # type: ignore[misc]
        except Exception:
            from dataclasses import replace
            r = replace(r, rank=new_rank)  # noqa: PLW2901
        canon.append(r)
    return canon



def write_submission(
    rows: Iterable[RankedCandidate],
    path: str | Path,
    *,
    strict: bool = True,
) -> Path:
    """Validate (when ``strict``) + write the top-K CSV.

    ``strict=True`` (the default) enforces every invariant the official
    ``validate_submission.py`` checks - this is what we use for the real
    100-row submission. ``strict=False`` is for debugging / smoke runs on
    small samples (e.g. the 50-row fixture) where we just want to inspect
    the shortlist locally.
    """
    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)

    materialised = list(rows)
    if strict:
        materialised = _enforce_invariants(materialised)
    else:
        materialised = sorted(materialised, key=lambda x: x.rank)

    with out.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh, quoting=csv.QUOTE_MINIMAL, lineterminator="\n")
        writer.writerow(HEADER)
        for r in materialised:
            # Clamp score to 4 decimals to keep the CSV diff-friendly.
            writer.writerow(
                [r.candidate_id, r.rank, f"{r.score:.4f}", r.reasoning.replace("\n", " ").strip()]
            )
    return out


def write_submission_xlsx(
    rows: Iterable[RankedCandidate],
    path: str | Path,
    *,
    strict: bool = True,
) -> Path:
    """Same as :func:`write_submission` but emits an .xlsx workbook.

    The hackathon submission spec accepts CSV or XLSX. XLSX is convenient
    for reviewers who want to open the shortlist in Excel without manually
    setting the text-import options.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - import guard
        raise SubmissionError(
            "openpyxl is required for XLSX submissions. Install with: pip install openpyxl"
        ) from exc

    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)

    materialised = list(rows)
    if strict:
        materialised = _enforce_invariants(materialised)
    else:
        materialised = sorted(materialised, key=lambda x: x.rank)

    wb = Workbook()
    ws = wb.active
    ws.title = "submission"

    # Header
    header_fill = PatternFill("solid", fgColor="111111")
    header_font = Font(name="Inter", bold=True, color="FFFFFF")
    for col, name in enumerate(HEADER, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"

    # Data
    for idx, r in enumerate(materialised, start=2):
        ws.cell(row=idx, column=1, value=r.candidate_id)
        ws.cell(row=idx, column=2, value=r.rank)
        ws.cell(row=idx, column=3, value=round(float(r.score), 4))
        ws.cell(row=idx, column=4, value=r.reasoning.replace("\n", " ").strip())

    # Column widths sized for readability
    widths = {1: 16, 2: 6, 3: 10, 4: 110}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=(cell.column == 4))

    wb.save(out)
    return out

